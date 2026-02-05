"""
Spreadsheet assignment evaluator for the school portal.
Evaluates student Excel submissions against an answer key and produces:
- Text/PDF report of where they went wrong
- Commented Excel (student file with cell comments)

Uses the same evaluator logic as the CTSS Spreadsheet evaluator if available;
otherwise uses the built-in evaluator (SALES_ANALYSIS mark scheme).
"""
import io
import os
import re
import sys
import tempfile
import logging
from pathlib import Path
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional, Dict, Any, List

logger = logging.getLogger(__name__)

# Optional: use external evaluator from CTSS project if path is set
SPREADSHEET_EVALUATOR_PATH = os.environ.get(
    'SPREADSHEET_EVALUATOR_PATH',
    str(Path(__file__).resolve().parents[2] / 'CTSS Class portal project' / 'Spreadsheet evaluator')
)


# ============== Built-in evaluator (used when external path is not available) ==============

@dataclass
class CellResult:
    """Result for a single cell evaluation"""
    cell_ref: str
    question_num: int
    marks_possible: int
    marks_awarded: float
    student_formula: Optional[str]
    expected_formula: Optional[str]
    student_value: Any
    expected_value: Any
    feedback: str
    formula_correct: bool
    value_correct: bool


@dataclass
class QuestionResult:
    """Result for a question (may span multiple cells)"""
    question_num: int
    description: str
    total_marks: int
    marks_awarded: float
    cells: List[CellResult] = field(default_factory=list)
    feedback: str = ""


@dataclass
class EvaluationResult:
    """Complete evaluation result"""
    student_file: str
    student_name: str
    total_marks: int
    marks_awarded: float
    percentage: float
    questions: List[QuestionResult] = field(default_factory=list)
    summary: str = ""


class MarkSchemeBuiltin:
    """Built-in mark scheme for SALES_ANALYSIS task (used when external evaluator unavailable)."""
    def __init__(self):
        self.questions = [
            {"num": 1, "description": "G4:G15 - Calculate 2025 Total Sales using SUM formula", "marks": 1,
             "cells": [f"G{i}" for i in range(4, 16)], "expected_formula_pattern": r"SUM\([C-F]\d+:[C-F]\d+\)",
             "formula_type": "SUM", "marking_notes": "1 mark for correct SUM formula summing quarterly sales"},
            {"num": 2, "description": "H4:H15 - Commission using VLOOKUP", "marks": 2,
             "cells": [f"H{i}" for i in range(4, 16)], "expected_formula_pattern": r"VLOOKUP\(.*\$?A\$?20.*\$?C\$?24.*3.*TRUE\).*\*",
             "formula_type": "VLOOKUP", "marking_notes": "1 mark VLOOKUP, 1 mark multiply"},
            {"num": 3, "description": "J4:J15 - IF Exceed/Miss", "marks": 2,
             "cells": [f"J{i}" for i in range(4, 16)], "expected_formula_pattern": r'IF\(G\d+>I\d+,.*EXCEED.*,.*MISS',
             "formula_type": "IF", "marking_notes": "1 mark IF, 1 mark logic"},
            {"num": 4, "description": "Conditional formatting - Red for miss", "marks": 3,
             "cells": ["A4:J15"], "expected_formula_pattern": None, "formula_type": "CONDITIONAL_FORMATTING",
             "marking_notes": "range, condition, red fill"},
            {"num": 5, "description": "I19:I22 - SUMIF departmental", "marks": 4,
             "cells": ["I19", "I20", "I21", "I22"], "expected_formula_pattern": r"SUMIF\(.*B.*:.*B.*,.*,.*G.*:.*G.*\)",
             "formula_type": "SUMIF", "marking_notes": "1 per department"},
            {"num": 6, "description": "I23 - SUM Company Total", "marks": 1,
             "cells": ["I23"], "expected_formula_pattern": r"SUM\(I19:I22\)", "formula_type": "SUM", "marking_notes": "SUM of departments"},
            {"num": 7, "description": "I25 - Sales yet to achieve", "marks": 1,
             "cells": ["I25"], "expected_formula_pattern": r"I24-I23", "formula_type": "SUBTRACTION", "marking_notes": "Target - Current"},
            {"num": 8, "description": "I26 - Days remaining", "marks": 1,
             "cells": ["I26"], "expected_formula_pattern": r"(DATEDIF|DAYS|H25-H23|DATE|82)", "formula_type": "DATE_CALCULATION", "marking_notes": "82 days"},
        ]

    def get_question(self, num: int) -> Optional[Dict]:
        for q in self.questions:
            if q["num"] == num:
                return q
        return None

    def get_total_marks(self) -> int:
        return sum(q["marks"] for q in self.questions)


class ExcelEvaluatorBuiltin:
    """Built-in Excel evaluator (same logic as CTSS evaluate_submissions)."""
    def __init__(self, answer_key_path: str, mark_scheme: MarkSchemeBuiltin = None):
        import openpyxl
        self.answer_key_path = answer_key_path
        self.mark_scheme = mark_scheme or MarkSchemeBuiltin()
        self.wb_ans = openpyxl.load_workbook(answer_key_path, data_only=False)
        self.ws_ans = self.wb_ans.active
        self.wb_ans_values = openpyxl.load_workbook(answer_key_path, data_only=True)
        self.ws_ans_values = self.wb_ans_values.active

    def normalize_formula(self, formula: str) -> str:
        if not formula:
            return ""
        f = formula.strip()
        if f.startswith("="):
            f = f[1:]
        f = f.upper()
        f = re.sub(r'\s+', '', f)
        return f

    def check_formula_pattern(self, formula: str, pattern: str) -> bool:
        if not formula or not pattern:
            return False
        try:
            return bool(re.search(pattern, self.normalize_formula(formula), re.IGNORECASE))
        except Exception:
            return False

    def compare_values(self, student_val: Any, expected_val: Any, tolerance: float = 0.01) -> bool:
        if student_val is None and expected_val is None:
            return True
        if student_val is None or expected_val is None:
            return False
        if isinstance(student_val, str) and isinstance(expected_val, str):
            return student_val.strip().lower() == expected_val.strip().lower()
        try:
            s_num = float(student_val)
            e_num = float(expected_val)
            if e_num == 0:
                return abs(s_num) < tolerance
            return abs(s_num - e_num) / abs(e_num) < tolerance
        except (ValueError, TypeError):
            pass
        return str(student_val).strip().lower() == str(expected_val).strip().lower()

    def evaluate_cell(self, ws_student, ws_student_values, cell_ref: str, question: Dict) -> CellResult:
        import openpyxl
        student_cell = ws_student[cell_ref]
        student_value_cell = ws_student_values[cell_ref]
        ans_cell = self.ws_ans[cell_ref]
        ans_value_cell = self.ws_ans_values[cell_ref]
        student_formula = student_cell.value if isinstance(student_cell.value, str) and (student_cell.value or '').startswith('=') else None
        expected_formula = ans_cell.value if isinstance(ans_cell.value, str) and (ans_cell.value or '').startswith('=') else None
        student_value = student_value_cell.value
        expected_value = ans_value_cell.value
        formula_correct = False
        if question.get("expected_formula_pattern") and student_formula:
            formula_correct = self.check_formula_pattern(student_formula, question["expected_formula_pattern"])
        value_correct = self.compare_values(student_value, expected_value)
        feedback_parts = []
        if not student_formula:
            feedback_parts.append("No formula entered")
        elif not formula_correct:
            feedback_parts.append(f"Formula structure incorrect. Expected pattern using {question.get('formula_type', 'formula')}")
        if not value_correct:
            feedback_parts.append(f"Value incorrect. Expected: {expected_value}, Got: {student_value}")
        if formula_correct and value_correct:
            feedback_parts.append("Correct!")
        return CellResult(
            cell_ref=cell_ref, question_num=question["num"], marks_possible=0, marks_awarded=0,
            student_formula=student_formula, expected_formula=expected_formula,
            student_value=student_value, expected_value=expected_value,
            feedback=" ".join(feedback_parts), formula_correct=formula_correct, value_correct=value_correct
        )

    def evaluate_conditional_formatting(self, ws_student, question: Dict) -> QuestionResult:
        result = QuestionResult(question_num=question["num"], description=question["description"], total_marks=question["marks"], marks_awarded=0)
        feedback_parts = []
        marks = 0
        cf_rules = getattr(ws_student.conditional_formatting, '_cf_rules', {}) or {}
        found_correct_range = False
        found_correct_formula = False
        found_red_fill = False
        for range_string, rules in cf_rules.items():
            range_str = str(range_string)
            if any(x in range_str.upper() for x in ['A4', 'J15', 'A4:J15']):
                found_correct_range = True
            for rule in rules:
                if getattr(rule, 'formula', None):
                    formula_str = str(rule.formula).upper()
                    if 'MISS' in formula_str or ('J' in formula_str and ('=' in formula_str or 'IF' in formula_str)):
                        found_correct_formula = True
                if getattr(rule, 'dxf', None) and getattr(rule.dxf, 'fill', None):
                    fill = rule.dxf.fill
                    if getattr(fill, 'bgColor', None) and fill.bgColor and getattr(fill.bgColor, 'rgb', None):
                        color = str(fill.bgColor.rgb).upper()
                        if len(color) >= 6:
                            try:
                                r = int(color[2:4], 16) if len(color) == 8 else int(color[0:2], 16)
                                g = int(color[4:6], 16) if len(color) == 8 else int(color[2:4], 16)
                                b = int(color[6:8], 16) if len(color) == 8 else int(color[4:6], 16)
                                if r > 200 and g < 150 and b < 150:
                                    found_red_fill = True
                            except Exception:
                                pass
        if found_correct_range:
            marks += 1
            feedback_parts.append("✓ Correct range applied.")
        else:
            feedback_parts.append("✗ Range should cover rows 4-15.")
        if found_correct_formula:
            marks += 1
            feedback_parts.append("✓ Condition formula correctly checks for 'Miss'.")
        else:
            feedback_parts.append("✗ Formula should check if column J = 'Miss'.")
        if found_red_fill:
            marks += 1
            feedback_parts.append("✓ Red background fill applied.")
        else:
            feedback_parts.append("✗ Fill color should be red.")
        result.marks_awarded = marks
        result.feedback = " ".join(feedback_parts)
        return result

    def evaluate_question(self, ws_student, ws_student_values, question: Dict) -> QuestionResult:
        if question.get("formula_type") == "CONDITIONAL_FORMATTING":
            return self.evaluate_conditional_formatting(ws_student, question)
        result = QuestionResult(question_num=question["num"], description=question["description"], total_marks=question["marks"], marks_awarded=0)
        for cell_ref in question["cells"]:
            result.cells.append(self.evaluate_cell(ws_student, ws_student_values, cell_ref, question))
        if question["num"] in [1, 6, 7, 8]:
            correct_cells = sum(1 for c in result.cells if c.formula_correct or c.value_correct)
            result.marks_awarded = question["marks"] if correct_cells == len(result.cells) else (0.5 if correct_cells > 0 else 0)
        elif question["num"] == 2:
            has_vlookup = any('VLOOKUP' in (c.student_formula or '').upper() for c in result.cells)
            has_multiply = any('*' in (c.student_formula or '') for c in result.cells)
            correct_values = sum(1 for c in result.cells if c.value_correct)
            result.marks_awarded = (1 if has_vlookup else 0) + (1 if (has_multiply and correct_values > len(result.cells) // 2) else 0)
        elif question["num"] == 3:
            has_if = any('IF(' in (c.student_formula or '').upper() for c in result.cells)
            correct_values = sum(1 for c in result.cells if c.value_correct)
            result.marks_awarded = (1 if has_if else 0) + (1 if correct_values == len(result.cells) else (0.5 if correct_values > len(result.cells) // 2 else 0))
        elif question["num"] == 5:
            result.marks_awarded = sum(1 for c in result.cells if c.formula_correct or c.value_correct)
        correct_count = sum(1 for c in result.cells if c.formula_correct and c.value_correct)
        total_count = len(result.cells)
        feedback_parts = [f"{correct_count}/{total_count} cells correct."]
        incorrect_cells = [c for c in result.cells if not (c.formula_correct and c.value_correct)]
        if incorrect_cells and len(incorrect_cells) <= 3:
            for c in incorrect_cells:
                feedback_parts.append(f"{c.cell_ref}: {c.feedback}")
        elif incorrect_cells:
            feedback_parts.append(f"Check cells: {', '.join(c.cell_ref for c in incorrect_cells[:5])}")
        result.feedback = " ".join(feedback_parts)
        return result

    def extract_student_name(self, filepath: str) -> str:
        filename = Path(filepath).stem
        match = re.search(r'SALES_ANALYSIS_(.+?)_\d+', filename, re.IGNORECASE)
        if match:
            return match.group(1).replace('_', ' ').title()
        clean_name = filename.replace('SALES_ANALYSIS_', '').replace('_', ' ')
        return clean_name if clean_name else "Unknown Student"

    def evaluate(self, student_file_path: str) -> EvaluationResult:
        import openpyxl
        wb_student = openpyxl.load_workbook(student_file_path, data_only=False)
        ws_student = wb_student.active
        wb_student_values = openpyxl.load_workbook(student_file_path, data_only=True)
        ws_student_values = wb_student_values.active
        student_name = self.extract_student_name(student_file_path)
        result = EvaluationResult(
            student_file=student_file_path, student_name=student_name,
            total_marks=self.mark_scheme.get_total_marks(), marks_awarded=0, percentage=0
        )
        for question in self.mark_scheme.questions:
            q_result = self.evaluate_question(ws_student, ws_student_values, question)
            result.questions.append(q_result)
            result.marks_awarded += q_result.marks_awarded
        result.percentage = (result.marks_awarded / result.total_marks) * 100 if result.total_marks > 0 else 0
        summary_parts = [f"Student: {student_name}", f"Total Score: {result.marks_awarded}/{result.total_marks} ({result.percentage:.1f}%)", "", "Question Breakdown:"]
        for q in result.questions:
            status = "✓" if q.marks_awarded == q.total_marks else "△" if q.marks_awarded > 0 else "✗"
            summary_parts.append(f"  Q{q.question_num}: {q.marks_awarded}/{q.total_marks} {status}")
        result.summary = "\n".join(summary_parts)
        return result


def _load_external_evaluator():
    """Import ExcelEvaluator and MarkScheme from external evaluator if available."""
    if SPREADSHEET_EVALUATOR_PATH and os.path.isdir(SPREADSHEET_EVALUATOR_PATH):
        if SPREADSHEET_EVALUATOR_PATH not in sys.path:
            sys.path.insert(0, SPREADSHEET_EVALUATOR_PATH)
        try:
            from evaluate_submissions import ExcelEvaluator, MarkScheme, generate_text_report as _generate_text_report
            return ExcelEvaluator, MarkScheme, _generate_text_report
        except ImportError as e:
            logger.warning(f"Could not import external spreadsheet evaluator: {e}")
    return None


def evaluate_spreadsheet_submission(
    answer_key_bytes: bytes,
    student_bytes: bytes,
    student_name: str = "Student",
    student_filename: str = "submission.xlsx",
) -> Optional[Dict[str, Any]]:
    """
    Evaluate a student Excel submission against the answer key.
    Returns a dict with marks_awarded, total_marks, percentage, questions (list of question results),
    summary (text), and full result for PDF/Excel generation; or None if evaluation fails.
    Uses external evaluator from SPREADSHEET_EVALUATOR_PATH if available; otherwise built-in evaluator.
    """
    loaded = _load_external_evaluator()
    if loaded:
        ExcelEvaluator, MarkScheme, _ = loaded
        evaluator_cls, scheme_cls = ExcelEvaluator, MarkScheme
    else:
        logger.info("Using built-in spreadsheet evaluator (SALES_ANALYSIS mark scheme)")
        evaluator_cls, scheme_cls = ExcelEvaluatorBuiltin, MarkSchemeBuiltin

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f_ans:
        f_ans.write(answer_key_bytes)
        ans_path = f_ans.name
    try:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f_stu:
            f_stu.write(student_bytes)
            stu_path = f_stu.name
        try:
            evaluator = evaluator_cls(ans_path, scheme_cls())
            result = evaluator.evaluate(stu_path)
            result.student_name = student_name
            result.student_file = student_filename
            return _result_to_dict(result)
        finally:
            try:
                os.unlink(stu_path)
            except Exception:
                pass
    finally:
        try:
            os.unlink(ans_path)
        except Exception:
            pass


def _result_to_dict(result) -> Dict[str, Any]:
    """Convert EvaluationResult to a JSON-serializable dict."""
    return {
        'student_name': result.student_name,
        'student_file': result.student_file,
        'total_marks': result.total_marks,
        'marks_awarded': result.marks_awarded,
        'percentage': result.percentage,
        'summary': result.summary,
        'questions': [
            {
                'question_num': q.question_num,
                'description': q.description,
                'total_marks': q.total_marks,
                'marks_awarded': q.marks_awarded,
                'feedback': q.feedback,
                'cells': [
                    {
                        'cell_ref': c.cell_ref,
                        'feedback': c.feedback,
                        'formula_correct': c.formula_correct,
                        'value_correct': c.value_correct,
                    }
                    for c in (q.cells or [])
                ]
            }
            for q in result.questions
        ],
    }


def generate_text_report(result_dict: Dict[str, Any]) -> str:
    """Generate plain text feedback report from result dict."""
    # Build text from dict (result_dict is JSON-serializable, no _result)
    lines = [
        "=" * 60,
        "EXCEL EVALUATION REPORT",
        "=" * 60,
        f"Student: {result_dict.get('student_name', 'Student')}",
        f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        f"TOTAL SCORE: {result_dict.get('marks_awarded', 0)}/{result_dict.get('total_marks', 0)} ({result_dict.get('percentage', 0):.1f}%)",
        "",
        "-" * 60,
        "QUESTION BREAKDOWN",
        "-" * 60,
    ]
    for q in result_dict.get('questions', []):
        status = "✓" if q.get('marks_awarded') == q.get('total_marks') else "△" if q.get('marks_awarded', 0) > 0 else "✗"
        lines.append(f"Q{q.get('question_num')}: {q.get('marks_awarded')}/{q.get('total_marks')} {status} - {q.get('description', '')}")
    lines.extend(["", "-" * 60, "DETAILED FEEDBACK", "-" * 60])
    for q in result_dict.get('questions', []):
        lines.append("")
        lines.append(f"Question {q.get('question_num')}: {q.get('description')}")
        lines.append(f"Marks: {q.get('marks_awarded')}/{q.get('total_marks')}")
        lines.append(f"Feedback: {q.get('feedback')}")
        for c in (q.get('cells') or [])[:5]:
            if not c.get('formula_correct') or not c.get('value_correct'):
                lines.append(f"  • {c.get('cell_ref')}: {c.get('feedback')}")
    lines.extend(["", "=" * 60, "END OF REPORT", "=" * 60])
    return "\n".join(lines)


def generate_pdf_report(result_dict: Dict[str, Any]) -> bytes:
    """Generate a PDF feedback report from the result dict."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
    from reportlab.lib.enums import TA_LEFT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=50, leftMargin=50, topMargin=50, bottomMargin=50)
    styles = getSampleStyleSheet()
    style = ParagraphStyle(
        name='Body',
        parent=styles['Normal'],
        fontSize=10,
        leading=14,
    )
    text = generate_text_report(result_dict)
    parts = []
    for line in text.splitlines():
        line = line.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
        if line.strip():
            parts.append(Paragraph(line, style))
        else:
            parts.append(Spacer(1, 12))
    doc.build(parts)
    return buf.getvalue()


def generate_commented_excel(student_bytes: bytes, result_dict: Dict[str, Any]) -> bytes:
    """Add feedback comments to the student's Excel file and return the workbook as bytes."""
    import openpyxl
    from openpyxl.comments import Comment

    wb = openpyxl.load_workbook(io.BytesIO(student_bytes))
    ws = wb.active
    author = "Feedback"

    for q in result_dict.get('questions', []):
        for c in (q.get('cells') or []):
            if not c.get('formula_correct') or not c.get('value_correct'):
                cell_ref = c.get('cell_ref')
                feedback = (c.get('feedback') or '').strip()
                if not cell_ref or not feedback:
                    continue
                try:
                    cell = ws[cell_ref]
                    comment_text = f"Q{q.get('question_num')}: {feedback}"
                    cell.comment = Comment(comment_text, author)
                except Exception as e:
                    logger.warning(f"Could not add comment to {cell_ref}: {e}")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
